# -*- coding: utf-8 -*-
"""
GERADOR DE CATALOGOS POR REPRESENTANTE - BRASIL BOTOES
=========================================================
O que este script faz:
  1. Le a planilha "ESTOQUES PROJETO V2 - Clientes" (aba "Cliente")
  2. Agrupa os produtos por REPRESENTANTE
  3. Para cada representante, gera uma pasta em representantes/<slug>/
     com um produtos.json (SO com os produtos daquele representante)
     e um index.html (copiado do template)
  4. Cruza com a tabela de cores (nome/familia/hex) e procura a foto
     de cada produto na pasta imagens/ do catalogo geral

IMPORTANTE: este catalogo e INDEPENDENTE do catalogo geral.
Ele nao le nem sobrescreve o produtos.json do catalogo geral -
os dados vem exclusivamente da planilha de Clientes.

REGRA DE FOTO - DIFERENTE DO CATALOGO GERAL:
No catalogo geral, produto sem foto NAO aparece no site.
Aqui NAO: todo produto vinculado a um representante entra no
produtos.json e no site, tenha foto ou nao (o card mostra "Sem foto").
Isso e proposital - o representante precisa ver o estoque completo
dos seus clientes, mesmo o que ainda nao foi fotografado.

COMO USAR:
  1. Ajuste as variaveis na secao CONFIGURACAO abaixo
  2. Rode: python gerar_catalogos_representantes.py
  3. Confira o relatorio_representantes.xlsx gerado (foto nao encontrada, etc.)
"""

import os
import re
import glob
import shutil
import unicodedata
from datetime import datetime

import openpyxl

# ======================= CONFIGURACAO =======================
# Planilha de Clientes (a que voce enviou / atualiza)
PLANILHA_CLIENTES = r"C:\Users\Usuario\BRASIL BOTOES LTDA\Estoque - Documentos\2026\ESTOQUES PROJETO V2 - Clientes.xlsm"
ABA_CLIENTES = "Cliente"
LINHA_CABECALHO = 11   # linha do cabecalho na planilha (a partir dela, +1 = dados)

# Pasta raiz do catalogo (onde fica o index.html geral, imagens/, etc.)
PASTA_CATALOGO = r"C:\Users\Usuario\BRASIL BOTOES LTDA\Público - Documentos\Catálogo Pronta Entrega"
# Caminhos confirmados em 16/08/2026 - se a planilha ou a pasta forem movidas, ajuste aqui.

# Tabela de cores (nome/familia/hex) - referencia compartilhada, nao e dado do catalogo geral
TABELA_CORES = os.path.join(PASTA_CATALOGO, "apoio", "tabela_cores_final.xlsx")

# Pasta de fotos (reaproveita as mesmas fotos do catalogo geral)
PASTA_IMAGENS = os.path.join(PASTA_CATALOGO, "imagens")

# Template do index.html do representante (fica dentro de documentacao/ ou apoio/)
TEMPLATE_INDEX = os.path.join(PASTA_CATALOGO, "apoio", "template_representante_index.html")

# Pasta onde os catalogos de cada representante serao gerados
PASTA_REPRESENTANTES = os.path.join(PASTA_CATALOGO, "representantes")

# Link do catalogo geral (para o botao "Ver catalogo geral" em cada pagina)
LINK_CATALOGO_GERAL = "https://brasilbotoes.github.io/catalogo"

# Apelidos curtos para a URL de cada representante (opcional, mas RECOMENDADO).
# Quem estiver aqui usa exatamente esse slug (fixo, nunca muda). Quem nao estiver
# aqui tem o slug abreviado automaticamente a partir do nome (ver abreviar_nome).
MAPA_SLUG = {
    "GABRIEL E PAPINI REPRESENTACOES LTDA": "gabrielepapini",
    "GRANDES CONTAS": "grandescontas",
    "J M PEREIRA E CIA LTDA.": "jmpereira",
    "E.P. REPRESENTACOES EIRELI": "eprepresentacoes",
    "STUDIO NEW REPRESENTACOES LTDA ME": "studionew",
    "A H DE OLIVEIRA REPRESENTACOES": "ahdeoliveira",
}

# Palavras que sao ignoradas ao abreviar o nome do representante (termos
# juridicos/genericos que nao ajudam a identificar a empresa na URL).
PALAVRAS_IGNORAR_SLUG = {
    "LTDA", "EIRELI", "ME", "EPP", "SA", "S/A", "CIA", "COMERCIO",
    "COMERCIAL", "REPRESENTACOES", "REPRESENTACAO", "REPRESENTACOES LTDA",
    "INDUSTRIA", "INDUSTRIAL", "DISTRIBUIDORA", "IMPORTACAO", "EXPORTACAO",
    "DE", "DO", "DA", "DOS", "DAS", "E",
}
# ==============================================================

EXTENSOES_VALIDAS = [".jpg", ".jpeg", ".png", ".webp"]


def slugify(texto):
    texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode()
    texto = re.sub(r"[^a-zA-Z0-9]+", "-", texto).strip("-").lower()
    return texto


def abreviar_nome(nome):
    """
    Reduz o nome completo do representante a palavra mais significativa,
    ignorando termos juridicos/genericos (LTDA, REPRESENTACOES, etc).
    Ex.: 'VISON REPRESENTACOES LTDA' -> 'vison'
         'MARVIC TEXTIL REPRESENTACOES LTDA' -> 'marvic'
    Se, depois de tirar essas palavras, nao sobrar nada, usa o nome inteiro.
    """
    texto = unicodedata.normalize("NFKD", nome).encode("ASCII", "ignore").decode()
    palavras = re.findall(r"[A-Za-z0-9]+", texto.upper())
    significativas = [p for p in palavras if p not in PALAVRAS_IGNORAR_SLUG]
    if not significativas:
        significativas = palavras
    return significativas[0].lower() if significativas else slugify(nome)


def slug_representante(nome):
    if nome in MAPA_SLUG:
        return MAPA_SLUG[nome]
    return abreviar_nome(nome)


def gerar_slugs_unicos(representantes):
    """
    Gera o slug de cada representante e resolve colisoes (dois representantes
    diferentes que abreviariam para o mesmo nome). Em caso de colisao, o
    segundo (e seguintes) ganham sufixo numerico: vison, vison-2, vison-3...
    Dica: se dois representantes colidirem com frequencia, cadastre um deles
    no MAPA_SLUG com um apelido fixo para nao depender da ordem da planilha.
    """
    usados = {}
    slugs = {}
    for rep in representantes:
        base = slug_representante(rep)
        slug = base
        contador = 2
        while slug in usados and usados[slug] != rep:
            slug = f"{base}-{contador}"
            contador += 1
        usados[slug] = rep
        slugs[rep] = slug
    return slugs


def cor_formatada(cor_raw):
    """Converte '0001' -> '001', '1993' -> '1993' (minimo 3 digitos, sem zeros a mais)."""
    try:
        return str(int(str(cor_raw).strip())).zfill(3)
    except (ValueError, TypeError):
        return str(cor_raw).strip()


def carregar_tabela_cores(caminho):
    cores = {}
    if not os.path.exists(caminho):
        print(f"AVISO: tabela de cores nao encontrada em {caminho} - seguindo sem nome/familia de cor.")
        return cores
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    headers = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h).strip() if h else "" for h in row]
        break
    idx = {h: i for i, h in enumerate(headers)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = row[idx.get("cor_codigo", 0)]
        if codigo is None:
            continue
        codigo_fmt = cor_formatada(codigo)
        cores[codigo_fmt] = {
            "nome": row[idx.get("cor_nome")] if "cor_nome" in idx else None,
            "familia": row[idx.get("cor_familia")] if "cor_familia" in idx else None,
            "hex": row[idx.get("cor_hex")] if "cor_hex" in idx else None,
        }
    return cores


def buscar_foto(modelo, cor_fmt, ting, furos):
    """Procura a foto principal do produto na pasta imagens/, no padrao do POP."""
    if not os.path.isdir(PASTA_IMAGENS):
        return None
    base = f"{modelo}_C{cor_fmt}"
    if ting:
        base += f"+{ting}"
    base += f"_{furos}"
    # foto principal = sem sufixo de angulo
    for ext in EXTENSOES_VALIDAS:
        candidato = os.path.join(PASTA_IMAGENS, base + ext)
        if os.path.exists(candidato):
            return os.path.relpath(candidato, PASTA_CATALOGO).replace("\\", "/")
    # busca case-insensitive / tolerante
    padrao = os.path.join(PASTA_IMAGENS, base + ".*")
    achados = glob.glob(padrao)
    if achados:
        return os.path.relpath(achados[0], PASTA_CATALOGO).replace("\\", "/")
    return None


def ler_produtos_por_representante():
    print("Lendo planilha de Clientes...")
    wb = openpyxl.load_workbook(PLANILHA_CLIENTES, read_only=True, data_only=True, keep_vba=False)
    ws = wb[ABA_CLIENTES]

    headers = None
    for row in ws.iter_rows(min_row=LINHA_CABECALHO, max_row=LINHA_CABECALHO, values_only=True):
        headers = [str(h).strip() if h else "" for h in row]
        break

    def col(nome):
        # pega o INDICE da 1a ocorrencia de uma coluna pelo nome exato
        return headers.index(nome) if nome in headers else None

    idx_codred = col("CÓD. RED.")
    idx_modelo = col("MODELO")
    idx_tamanho = col("TAMANHO")
    idx_cor = col("COR")
    idx_ting = col("TING.")
    idx_desccor = col("DESC. COR")
    idx_furos = col("FUROS")
    idx_acabam = col("ACABAM.")
    idx_material = col("MATERIAL")
    idx_descr = col("DESCR.")
    idx_qtd = col("QTD SALDO")
    idx_um = col("UM")
    idx_cliente = col("CLIENTE")
    idx_rep = col("REPRESENTANTE")
    idx_codlaser = col("CÓD. LASER")
    idx_desclaser = col("DESC")

    if idx_codlaser is None or idx_desclaser is None:
        print("AVISO: coluna 'CÓD. LASER' e/ou 'DESC' nao encontrada na aba Cliente - "
              "o codigo de personalizacao nao sera exibido no catalogo do representante. "
              "Confira se o nome da(s) coluna(s) na planilha bate exatamente com esses nomes.")
    if idx_um is None:
        print("AVISO: coluna 'UM' nao encontrada na aba Cliente - "
              "o saldo sera exibido sem unidade de medida.")

    faltando = [n for n, i in [
        ("CÓD. RED.", idx_codred), ("MODELO", idx_modelo), ("COR", idx_cor),
        ("FUROS", idx_furos), ("CLIENTE", idx_cliente), ("REPRESENTANTE", idx_rep)
    ] if i is None]
    if faltando:
        raise ValueError(f"Colunas nao encontradas na linha {LINHA_CABECALHO}: {faltando}. "
                          f"Confira LINHA_CABECALHO na configuracao.")

    cores = carregar_tabela_cores(TABELA_CORES)

    produtos_por_rep = {}
    relatorio = []

    for row in ws.iter_rows(min_row=LINHA_CABECALHO + 1, values_only=True):
        codred = row[idx_codred]
        rep = row[idx_rep]
        if codred is None or not rep:
            continue  # sem representante, nao entra em nenhum catalogo

        modelo = str(row[idx_modelo]).strip() if row[idx_modelo] is not None else ""
        cor_raw = row[idx_cor]
        cor_fmt = cor_formatada(cor_raw) if cor_raw is not None else ""
        ting_raw = row[idx_ting] if idx_ting is not None else None
        ting = str(ting_raw).strip() if ting_raw not in (None, "", "SIM") else None
        furos = str(row[idx_furos]).strip() if row[idx_furos] is not None else ""

        foto = buscar_foto(modelo, cor_fmt, ting, furos)
        # Sem foto? Segue no catalogo mesmo assim (regra diferente do geral - ver topo do arquivo).

        # Codigo de personalizacao: "CÓD. LASER - DESC." (ex.: "P2560 - GATA BAKANA")
        codlaser = str(row[idx_codlaser]).strip() if idx_codlaser is not None and row[idx_codlaser] not in (None, "") else None
        desclaser = str(row[idx_desclaser]).strip() if idx_desclaser is not None and row[idx_desclaser] not in (None, "") else None
        if codlaser and desclaser:
            personalizacao = f"{codlaser} - {desclaser}"
        elif codlaser:
            personalizacao = codlaser
        else:
            personalizacao = None

        info_cor = cores.get(cor_fmt, {})

        produto = {
            "cod_red": str(codred),
            "modelo": modelo,
            "tamanho": str(row[idx_tamanho]).strip() if idx_tamanho is not None and row[idx_tamanho] is not None else None,
            "cor_codigo": f"C{cor_fmt}",
            "tingimento": ting,
            "personalizacao": personalizacao,
            "cor_nome": (row[idx_desccor] if idx_desccor is not None and row[idx_desccor] else info_cor.get("nome")),
            "cor_familia": info_cor.get("familia"),
            "cor_hex": info_cor.get("hex"),
            "furacao": furos,
            "acabamento": row[idx_acabam] if idx_acabam is not None else None,
            "material": row[idx_material] if idx_material is not None else None,
            "descricao": row[idx_descr] if idx_descr is not None else None,
            "qtd_saldo": row[idx_qtd] if idx_qtd is not None else None,
            "unidade": row[idx_um] if idx_um is not None and row[idx_um] not in (None, "") else None,
            "cliente": row[idx_cliente],
            "foto": foto,
        }

        produtos_por_rep.setdefault(rep, []).append(produto)

        relatorio.append({
            "representante": rep,
            "cliente": produto["cliente"],
            "cod_red": produto["cod_red"],
            "modelo": modelo,
            "cor": produto["cor_codigo"],
            "foto_encontrada": "SIM" if foto else "NAO",
        })

    return produtos_por_rep, relatorio


def gerar_paginas(produtos_por_rep):
    if not os.path.exists(TEMPLATE_INDEX):
        print(f"AVISO: template nao encontrado em {TEMPLATE_INDEX}. "
              f"Os produtos.json foram gerados, mas o index.html de cada representante precisa ser criado manualmente.")
        template_html = None
    else:
        with open(TEMPLATE_INDEX, "r", encoding="utf-8") as f:
            template_html = f.read()

    os.makedirs(PASTA_REPRESENTANTES, exist_ok=True)
    links = []

    slugs = gerar_slugs_unicos(produtos_por_rep.keys())

    for rep, produtos in produtos_por_rep.items():
        slug = slugs[rep]
        pasta_rep = os.path.join(PASTA_REPRESENTANTES, slug)
        os.makedirs(pasta_rep, exist_ok=True)

        # produtos.json
        import json
        with open(os.path.join(pasta_rep, "produtos.json"), "w", encoding="utf-8") as f:
            json.dump(produtos, f, ensure_ascii=False, indent=2)

        # index.html (a partir do template, com nome do representante substituido)
        if template_html:
            html = template_html.replace("{{REPRESENTANTE}}", rep)
            html = html.replace("{{LINK_CATALOGO_GERAL}}", LINK_CATALOGO_GERAL)
            with open(os.path.join(pasta_rep, "index.html"), "w", encoding="utf-8") as f:
                f.write(html)

        links.append((rep, slug, len(produtos)))
        print(f"  -> {rep}: {len(produtos)} produtos -> representantes/{slug}/")

    return links


def gerar_relatorio(relatorio, links):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos por Representante"
    ws.append(["Representante", "Cliente", "Cód. Red.", "Modelo", "Cor", "Foto encontrada"])
    for linha in relatorio:
        ws.append([linha["representante"], linha["cliente"], linha["cod_red"],
                   linha["modelo"], linha["cor"], linha["foto_encontrada"]])
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 30

    ws2 = wb.create_sheet("Resumo")
    ws2.append(["Representante", "Slug (URL)", "Qtd. Produtos"])
    for rep, slug, qtd in links:
        ws2.append([rep, slug, qtd])
    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 35

    nome = f"relatorio_representantes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), nome)
    wb.save(caminho)
    print(f"\nRelatorio salvo em: {caminho}")


def main():
    produtos_por_rep, relatorio = ler_produtos_por_representante()
    print(f"\n{len(produtos_por_rep)} representantes encontrados.\n")
    links = gerar_paginas(produtos_por_rep)
    sem_foto = sum(1 for l in relatorio if l["foto_encontrada"] == "NAO")
    print(f"\nTotal de produtos processados: {len(relatorio)}")
    print(f"Sem foto encontrada: {sem_foto}")
    gerar_relatorio(relatorio, links)


if __name__ == "__main__":
    main()
