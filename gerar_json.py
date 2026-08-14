"""
gerar_json.py — Catálogo Digital Brasil Botões
Lê ESTOQUES_PROJETO_V2_-_Saldos2.xlsm e tabela_cores_final.xlsx
Gera produtos.json e copia imagens de 'Itens Classificados' para 'imagens/'
"""

import json
import os
import re
import shutil
from pathlib import Path

import openpyxl

# ─── CAMINHOS ─────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent  # pasta onde está este script (catálogo)

PLANILHA = Path(r"C:\Users\Usuario\BRASIL BOTOES LTDA\Estoque - Documentos\2026\ESTOQUES PROJETO V2 - Saldos2.xlsm")
TABELA_CORES = BASE_DIR / "apoio" / "tabela_cores_final.xlsx"
PASTA_FOTOS_ORIGEM = Path(r"C:\Users\Usuario\BRASIL BOTOES LTDA\Produtos - Documentos\Itens Classificados")
PASTA_IMAGENS = BASE_DIR / "imagens"
ARQUIVO_JSON = BASE_DIR / "produtos.json"

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def normalizar_cor(codigo):
    """Normaliza código de cor: '0001'→'C001', 'C164'→'C164', '164'→'C164'"""
    codigo = str(codigo).strip()
    if codigo.upper().startswith("C"):
        return codigo.upper()
    codigo_num = codigo.lstrip("0") or "0"
    try:
        return f"C{int(codigo_num):03d}"
    except ValueError:
        return f"C{codigo}"

def montar_prefixo_imagem(modelo, cor_codigo, ting):
    """
    Monta o prefixo do nome do arquivo de imagem.
    Sem tingimento:  '17_C001'
    Com tingimento:  '17_C160+134'
    """
    modelo = str(modelo).strip()
    cor = normalizar_cor(cor_codigo)
    ting = str(ting).strip() if ting else ""
    if ting:
        return f"{modelo}_{cor}+{ting}"
    return f"{modelo}_{cor}"

def buscar_imagens(prefixo_com_furacao):
    """
    Varre PASTA_FOTOS_ORIGEM buscando arquivos que começam com o prefixo.
    Copia para PASTA_IMAGENS e retorna lista de nomes de arquivo.
    """
    encontrados = []
    if not PASTA_FOTOS_ORIGEM.exists():
        return encontrados

    prefixo_lower = prefixo_com_furacao.lower()
    for arquivo in PASTA_FOTOS_ORIGEM.iterdir():
        if not arquivo.is_file():
            continue
        nome = arquivo.stem.lower()  # sem extensão
        ext = arquivo.suffix.lower()
        if ext not in (".jpg", ".jpeg", ".png", ".webp"):
            continue
        # Verifica se o nome começa com o prefixo (case-insensitive)
        if nome == prefixo_lower or nome.startswith(prefixo_lower + "_"):
            destino = PASTA_IMAGENS / arquivo.name
            if not destino.exists():
                shutil.copy2(arquivo, destino)
            encontrados.append(arquivo.name)

    # Ordena: foto principal primeiro, depois variações
    def ordem(nome):
        n = nome.lower()
        if "_frente" in n:
            return 1
        if "_verso" in n:
            return 2
        if "_lado" in n or "_lat" in n:
            return 3
        if "_detalhe" in n:
            return 4
        return 0  # foto principal (sem sufixo)

    encontrados.sort(key=ordem)
    return encontrados

# ─── CARREGA TABELA DE CORES ──────────────────────────────────────────────────

def carregar_tabela_cores():
    """Retorna dict: {'C001': {'nome': ..., 'familia': ..., 'hex': ...}}"""
    tabela = {}
    if not TABELA_CORES.exists():
        print(f"  [AVISO] tabela_cores_final.xlsx não encontrada em {TABELA_CORES}")
        return tabela
    wb = openpyxl.load_workbook(TABELA_CORES, read_only=True, data_only=True)
    ws = wb.active
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip().lower() if h else "" for h in row]
            continue
        if not row[0]:
            continue
        linha = dict(zip(headers, row))
        codigo = str(linha.get("cor_codigo", "")).strip().upper()
        if codigo:
            tabela[codigo] = {
                "nome": str(linha.get("cor_nome", "") or "").strip(),
                "familia": str(linha.get("cor_familia", "") or "").strip(),
                "hex": str(linha.get("cor_hex", "") or "").strip(),
            }
    return tabela

# ─── CARREGA CADPROD ──────────────────────────────────────────────────────────

def carregar_cadprod(wb):
    """Retorna dict: {cod_produto: {descr, modelo, tamanho, cor, ting, furos, acabam}}"""
    ws = wb["CadProd"]
    headers = None
    dados = {}
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip().lower() if h else "" for h in row]
            continue
        # Ignora linhas sem código (primeira coluna)
        if row[0] is None or str(row[0]).strip() == "":
            continue
        linha = dict(zip(headers, row))
        # Tenta as variações possíveis do nome da coluna de código
        cod = None
        for chave in ("codproduto", "cod. produto", "código", "codigo"):
            val = linha.get(chave)
            if val is not None:
                cod = str(val).strip()
                break
        if not cod:
            # Usa a primeira coluna diretamente
            cod = str(row[0]).strip()
        if cod:
            dados[cod] = {
                "descr": str(linha.get("descr.", "") or linha.get("descr", "") or "").strip(),
                "modelo": str(linha.get("modelo", "") or "").strip(),
                "tamanho": str(linha.get("tamanho", "") or "").strip(),
                "cor": str(linha.get("cor", "") or "").strip(),
                "ting": str(linha.get("ting.", "") or linha.get("ting", "") or "").strip(),
                "furos": str(linha.get("furos", "") or "").strip(),
                "acabam": str(linha.get("acabam.", "") or linha.get("acabam", "") or "").strip(),
            }
    return dados

# ─── PRINCIPAL ────────────────────────────────────────────────────────────────

def main():
    print("=" * 52)
    print("  Brasil Botões — Gerando catálogo")
    print("=" * 52)

    # Verifica planilha
    if not PLANILHA.exists():
        print(f"\n[ERRO] Planilha não encontrada:\n  {PLANILHA}")
        raise SystemExit(1)

    print(f"\n[1/4] Carregando planilha...")
    wb = openpyxl.load_workbook(PLANILHA, read_only=True, keep_vba=True, data_only=True)

    tabela_cores = carregar_tabela_cores()
    print(f"       {len(tabela_cores)} cores na tabela de referência")

    cadprod = carregar_cadprod(wb)
    print(f"       {len(cadprod)} produtos no cadastro (CadProd)")

    # Garante pasta imagens
    PASTA_IMAGENS.mkdir(exist_ok=True)

    # Lê aba Estoque
    print(f"\n[2/4] Lendo saldos (aba Estoque)...")
    ws_est = wb["Estoque"]
    headers = None
    header_row = None

    for row in ws_est.iter_rows(values_only=True):
        if row[0] == "SEQ.":
            headers = [str(h).strip() if h else "" for h in row]
            header_row = True
            continue
        if not header_row or not row[0]:
            continue
        # A partir daqui são dados reais
        break

    # Re-lê do início para processar todas as linhas de dados
    headers = None
    dados_encontrados = False
    produtos = []
    imagens_copiadas = 0
    inativos_ignorados = 0
    sem_foto_ignorados = 0

    for row in ws_est.iter_rows(values_only=True):
        if not dados_encontrados:
            if row[0] == "SEQ.":
                headers = [str(h).strip() if h else "" for h in row]
                dados_encontrados = True
            continue

        # Linha de dados: SEQ. deve ser número
        if not row[0] or not str(row[0]).strip().isdigit():
            continue

        linha = dict(zip(headers, row))

        cod_red = str(linha.get("CÓD. RED.", "") or "").strip()
        modelo  = str(linha.get("MODELO", "") or "").strip()
        tamanho = str(linha.get("TAMANHO", "") or "").strip()
        cor_raw = str(linha.get("COR", "") or "").strip()
        ting    = str(linha.get("TING.", "") or "").strip()
        desc_cor= str(linha.get("DESC. COR", "") or "").strip()
        furos   = str(linha.get("FUROS", "") or "").strip()
        acabam  = str(linha.get("ACABAM.", "") or "").strip()
        obs     = str(linha.get("OBS", "") or "").strip()
        material= str(linha.get("MATERIAL", "") or "").strip()
        situacao= str(linha.get("SITUAÇÃO", "") or linha.get("SITUACAO", "") or "").strip()

        # Só inclui produtos com situação Ativo (ignora Inativo, vazio, etc.)
        if situacao.strip().lower() != "ativo":
            inativos_ignorados += 1
            continue

        saldo   = linha.get("QTD SALDO")
        um      = str(linha.get("UM", "") or "").strip()
        descr   = str(linha.get("DESCR.", "") or "").strip()

        # Quantidade
        try:
            qtd = float(saldo) if saldo is not None else 0
        except (TypeError, ValueError):
            qtd = 0

        # Código de cor normalizado
        cor_codigo = normalizar_cor(cor_raw) if cor_raw else ""

        # Nome da cor: DESC. COR da planilha tem prioridade, senão tabela
        if desc_cor:
            cor_nome = desc_cor
            cor_familia = tabela_cores.get(cor_codigo, {}).get("familia", "")
            cor_hex = tabela_cores.get(cor_codigo, {}).get("hex", "")
        else:
            ref_tabela = tabela_cores.get(cor_codigo, {})
            cor_nome = ref_tabela.get("nome", "")
            cor_familia = ref_tabela.get("familia", "")
            cor_hex = ref_tabela.get("hex", "")

        # Monta prefixo e busca imagens
        prefixo = montar_prefixo_imagem(modelo, cor_raw, ting)
        prefixo_furacao = f"{prefixo}_{furos}" if furos else prefixo
        imagens = buscar_imagens(prefixo_furacao)
        if imagens:
            imagens_copiadas += len(imagens)
        else:
            sem_foto_ignorados += 1
            continue  # só entram no catálogo produtos com pelo menos 1 foto

        # imagem_base: prefixo usado pelo index.html para carregar imagens
        imagem_base = prefixo_furacao  # ex: 17_C001_4F ou 17_C160+134_4F

        produto = {
            "reduzido": cod_red,        # código reduzido
            "referencia": modelo,       # modelo = referência no catálogo
            "descricao": descr,
            "tamanho": tamanho,
            "cor_codigo": cor_codigo,
            "cor_nome": cor_nome,
            "cor_familia": cor_familia,
            "cor_hex": cor_hex,
            "tingimento": ting,
            "furacao": furos,
            "acabamento": acabam,
            "material": material,
            "obs": obs,
            "saldo": qtd,
            "unidade": um,
            "imagem_base": imagem_base, # usado pelo index.html para carregar imagens
            "imagens": imagens,
        }
        produtos.append(produto)

    print(f"       {len(produtos)} produtos com saldo lidos (Ativo)")
    print(f"       {inativos_ignorados} produto(s) ignorado(s) por não estarem 'Ativo'")
    print(f"       {sem_foto_ignorados} produto(s) ignorado(s) por não terem foto")
    print(f"\n[3/4] Imagens copiadas para pasta imagens/: {imagens_copiadas} arquivo(s)")

    # Salva JSON
    print(f"\n[4/4] Gravando produtos.json...")
    with open(ARQUIVO_JSON, "w", encoding="utf-8") as f:
        json.dump(produtos, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*52}")
    print(f"  Concluído! {len(produtos)} produtos exportados.")
    print(f"{'='*52}\n")

if __name__ == "__main__":
    main()
