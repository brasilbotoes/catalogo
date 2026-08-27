# -*- coding: utf-8 -*-
"""
VALIDADOR DE FOTOS POR CODIGO DE COR - BRASIL BOTOES
=======================================================
O que este script faz:
  1. Le todos os cor_codigo da tabela_cores_final.xlsx
  2. Varre a pasta de fotos (Itens Classificados) e confere quais codigos
     aparecem em pelo menos um nome de arquivo
  3. Gera um relatorio Excel dizendo, pra cada codigo: TEM FOTO ou NAO TEM,
     e qual foi o arquivo encontrado (quando tem)

COMO USAR:
  1. Ajuste as 2 variaveis na secao "CONFIGURACAO" abaixo
  2. Rode: python validar_fotos_cores.py
  3. Abra o relatorio gerado (relatorio_fotos_cores.xlsx)
"""

import os
import re
from datetime import datetime

import openpyxl

# ======================= CONFIGURACAO =======================
# Pasta onde estao as fotos ja nomeadas (padrao do POP)
PASTA_FOTOS = r"C:\Users\Usuario\BRASIL BOTOES LTDA\Produtos - Documentos\Itens Classificados"

# Planilha de cores atualizada
ARQUIVO_TABELA = r"tabela_cores_final.xlsx"
ABA_TABELA = "Cores"
# ==============================================================

EXTENSOES_VALIDAS = {".jpg", ".jpeg", ".png", ".webp"}


def carregar_codigos(caminho_tabela, aba):
    """Le a coluna cor_codigo e cor_nome da tabela de cores."""
    wb = openpyxl.load_workbook(caminho_tabela, data_only=True)
    ws = wb[aba]

    headers = [c.value for c in ws[1]]
    idx_codigo = headers.index("cor_codigo")
    idx_nome = headers.index("cor_nome")
    idx_imagem = headers.index("imagem_origem")

    codigos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = row[idx_codigo]
        if not codigo:
            continue
        codigos.append({
            "cor_codigo": str(codigo).strip(),
            "cor_nome": row[idx_nome],
            "ja_tinha_imagem_origem": bool(row[idx_imagem]),
        })
    return codigos


def listar_arquivos_fotos(pasta):
    """Varre a pasta (e subpastas) e retorna a lista de nomes de arquivo (sem extensao)."""
    arquivos = []
    for raiz, _subpastas, nomes in os.walk(pasta):
        for nome in nomes:
            ext = os.path.splitext(nome)[1].lower()
            if ext in EXTENSOES_VALIDAS:
                arquivos.append(nome)
    return arquivos


def encontrar_foto(codigo, arquivos):
    """
    Procura o codigo como um TOKEN isolado no nome do arquivo
    (delimitado por _ , . ou inicio/fim), pra evitar falso positivo
    tipo 'C001' bater dentro de 'C0011'.
    """
    padrao = re.compile(
        r'(?:^|_)' + re.escape(codigo) + r'(?:_|\.)',
        re.IGNORECASE
    )
    for nome_arquivo in arquivos:
        if padrao.search(nome_arquivo):
            return nome_arquivo
    return None


def main():
    print("Lendo tabela de cores...")
    codigos = carregar_codigos(ARQUIVO_TABELA, ABA_TABELA)
    print(f"  -> {len(codigos)} codigos de cor na tabela")

    print("Varrendo pasta de fotos...")
    arquivos = listar_arquivos_fotos(PASTA_FOTOS)
    print(f"  -> {len(arquivos)} imagens encontradas na pasta")

    print("Cruzando codigos com arquivos...")
    linhas_relatorio = []
    tem_foto_count = 0
    for item in codigos:
        arquivo_encontrado = encontrar_foto(item["cor_codigo"], arquivos)
        tem_foto = arquivo_encontrado is not None
        if tem_foto:
            tem_foto_count += 1
        linhas_relatorio.append({
            "cor_codigo": item["cor_codigo"],
            "cor_nome": item["cor_nome"],
            "tem_foto": "SIM" if tem_foto else "NAO",
            "arquivo_encontrado": arquivo_encontrado or "",
            "ja_tinha_imagem_origem_antes": "SIM" if item["ja_tinha_imagem_origem"] else "NAO",
        })

    # --- Gera relatorio em Excel ---
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Relatorio"
    ws_out.append(["cor_codigo", "cor_nome", "tem_foto", "arquivo_encontrado", "ja_tinha_imagem_origem_antes"])
    for linha in linhas_relatorio:
        ws_out.append([
            linha["cor_codigo"],
            linha["cor_nome"],
            linha["tem_foto"],
            linha["arquivo_encontrado"],
            linha["ja_tinha_imagem_origem_antes"],
        ])
    for col in ["A", "B", "C", "D", "E"]:
        ws_out.column_dimensions[col].width = 30

    nome_relatorio = f"relatorio_fotos_cores_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    caminho_relatorio = os.path.join(os.path.dirname(os.path.abspath(__file__)), nome_relatorio)
    wb_out.save(caminho_relatorio)

    print()
    print("=" * 50)
    print(f"Total de codigos de cor: {len(codigos)}")
    print(f"Com foto encontrada: {tem_foto_count}")
    print(f"Sem foto encontrada: {len(codigos) - tem_foto_count}")
    print(f"Relatorio salvo em: {caminho_relatorio}")
    print("=" * 50)


if __name__ == "__main__":
    main()
